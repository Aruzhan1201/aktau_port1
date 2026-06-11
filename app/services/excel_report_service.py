import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.cargo import Cargo, CargoStatus
from app.models.berth import Berth, BerthStatus
from app.models.berth_reservation import BerthReservation
from app.models.parking_spot import ParkingSpot, ParkingSpotStatus
from app.models.parking_zone import ParkingZone
from app.models.payment import Payment, PaymentStatus
from app.models.port_queue import PortQueue, QueueStatus
from app.models.incident_report import IncidentReport
from app.models.ro_ro_vehicle import RoRoVehicle
from app.models.deal import Deal, DealStatus


async def generate_excel_report(
    session: AsyncSession,
    date_from: datetime,
    date_to: datetime,
) -> io.BytesIO:
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_width(ws, num_cols):
        for col in range(1, num_cols + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 50)

    # --- Sheet 1: Cargo / Transit Flow ---
    ws1 = wb.active
    ws1.title = "Transit Flow"
    ws1.append(["ID", "Cargo Type", "Weight", "Origin", "Destination", "Status", "Vehicle Type", "Budget",
                "Client ID", "Driver ID", "Ship ID", "Created At", "Updated At"])
    result = await session.execute(
        select(Cargo).where(Cargo.created_at.between(date_from, date_to))
    )
    for c in result.scalars().all():
        ws1.append([c.id, c.cargo_type, c.weight, c.origin, c.destination,
                    c.status.value, c.vehicle_type.value if c.vehicle_type else "",
                    c.budget, c.client_id, c.driver_id, c.ship_id,
                    c.created_at.isoformat() if c.created_at else "",
                    c.updated_at.isoformat() if c.updated_at else ""])
    style_header(ws1, 13)
    auto_width(ws1, 13)

    # --- Sheet 2: Payments / Money Flow ---
    ws2 = wb.create_sheet("Money Flow")
    ws2.append(["ID", "Type", "Amount", "Currency", "Status", "Cargo ID", "Reservation ID", "Paid By", "Created At", "Paid At"])
    result = await session.execute(
        select(Payment).where(Payment.created_at.between(date_from, date_to))
    )
    for p in result.scalars().all():
        ws2.append([p.id, p.type.value if hasattr(p.type, "value") else p.type,
                    p.amount, p.currency,
                    p.status.value if hasattr(p.status, "value") else p.status,
                    p.cargo_id, p.reservation_id, p.paid_by,
                    p.created_at.isoformat() if p.created_at else "",
                    p.paid_at.isoformat() if p.paid_at else ""])
    style_header(ws2, 10)
    auto_width(ws2, 10)

    # --- Sheet 3: Berth Occupancy ---
    ws3 = wb.create_sheet("Berth Occupancy")
    ws3.append(["ID", "Berth Name", "Status", "Capacity", "Manager ID", "Created At"])
    result = await session.execute(select(Berth))
    for b in result.scalars().all():
        ws3.append([b.id, b.name, b.status.value if hasattr(b.status, "value") else b.status,
                    b.capacity, b.manager_id,
                    b.created_at.isoformat() if b.created_at else ""])
    style_header(ws3, 6)
    auto_width(ws3, 6)

    # --- Sheet 4: Reservations ---
    ws4 = wb.create_sheet("Reservations")
    ws4.append(["ID", "Berth ID", "Ship ID", "Status", "Arrival", "Departure", "Created At"])
    result = await session.execute(
        select(BerthReservation).where(BerthReservation.created_at.between(date_from, date_to))
    )
    for r in result.scalars().all():
        ws4.append([r.id, r.berth_id, r.ship_id,
                    r.status.value if hasattr(r.status, "value") else r.status,
                    r.arrival_time.isoformat() if r.arrival_time else "",
                    r.departure_time.isoformat() if r.departure_time else "",
                    r.created_at.isoformat() if r.created_at else ""])
    style_header(ws4, 7)
    auto_width(ws4, 7)

    # --- Sheet 5: Parking ---
    ws5 = wb.create_sheet("Parking")
    ws5.append(["Spot ID", "Zone", "Spot Number", "Status", "Driver ID", "Tariff/hr", "Time In", "Time Out"])
    result = await session.execute(
        select(ParkingSpot, ParkingZone.name).join(ParkingZone, ParkingSpot.zone_id == ParkingZone.id)
    )
    for spot, zone_name in result.all():
        ws5.append([spot.id, zone_name, spot.spot_number,
                    spot.status.value if hasattr(spot.status, "value") else spot.status,
                    spot.driver_id, spot.tariff_per_hour,
                    spot.time_in.isoformat() if spot.time_in else "",
                    spot.time_out.isoformat() if spot.time_out else ""])
    style_header(ws5, 8)
    auto_width(ws5, 8)

    # --- Sheet 6: Deals ---
    ws6 = wb.create_sheet("Deals")
    ws6.append(["ID", "Type", "Status", "Client ID", "Driver ID", "Captain ID",
                "Cargo ID", "Price", "Currency", "Created At"])
    result = await session.execute(
        select(Deal).where(Deal.created_at.between(date_from, date_to))
    )
    for d in result.scalars().all():
        ws6.append([d.id, d.type.value if hasattr(d.type, "value") else d.type,
                    d.status.value if hasattr(d.status, "value") else d.status,
                    d.client_id, d.driver_id, d.captain_id,
                    d.cargo_id, d.proposed_price, d.currency,
                    d.created_at.isoformat() if d.created_at else ""])
    style_header(ws6, 10)
    auto_width(ws6, 10)

    # --- Sheet 7: Incidents ---
    ws7 = wb.create_sheet("Incidents")
    ws7.append(["ID", "Port", "Type", "Severity", "Status", "Reported By", "Created At"])
    result = await session.execute(
        select(IncidentReport).where(IncidentReport.created_at.between(date_from, date_to))
    )
    for i in result.scalars().all():
        ws7.append([i.id, i.port, i.incident_type,
                    i.severity.value if hasattr(i.severity, "value") else i.severity,
                    i.status.value if hasattr(i.status, "value") else i.status,
                    i.reported_by,
                    i.created_at.isoformat() if i.created_at else ""])
    style_header(ws7, 7)
    auto_width(ws7, 7)

    # --- Sheet 8: Ro-Ro ---
    ws8 = wb.create_sheet("Ro-Ro Vehicles")
    ws8.append(["ID", "Plate", "Driver", "Type", "Port", "Status", "Entry Time", "Exit Time"])
    result = await session.execute(
        select(RoRoVehicle).where(RoRoVehicle.entry_time.between(date_from, date_to))
    )
    for v in result.scalars().all():
        ws8.append([v.id, v.plate_number, v.driver_name, v.vehicle_type, v.port,
                    v.status.value if hasattr(v.status, "value") else v.status,
                    v.entry_time.isoformat() if v.entry_time else "",
                    v.exit_time.isoformat() if v.exit_time else ""])
    style_header(ws8, 8)
    auto_width(ws8, 8)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
